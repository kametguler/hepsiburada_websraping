import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from utils import consts, scroll_to_element, hover_to_element
from selenium.webdriver.chrome.options import Options
from selenium.common import exceptions


class Scraper:
    founded_count = 0
    product_per_page = 24

    def __init__(self, keyword, count):
        options = Options()
        # options.headless = True
        self.browser = webdriver.Chrome(executable_path=consts.CHROME_DRIVER_PATH, options=options)
        self.keyword = keyword
        self.count = count
        self.search_url = self.make_search_url()
        self.browser.get(self.search_url)
        self.wait = WebDriverWait(self.browser, 15)
        self.products_list = []

    def make_search_url(self):
        return f"{consts.SITE_URL}ara?q={self.keyword.replace(' ', '+')}"

    def check_enough_or_not(self):
        try:
            text = self.wait.until(EC.presence_of_element_located((By.XPATH,
                                                                   consts.XPATH_FOUNDED_PRODUCT_COUNT.replace(
                                                                       "@keyword", self.keyword)))).text
        except:
            raise exceptions.NoSuchElementException("Aranan anahtar kelimelerine göre ürün bulunamadı.")

        if "+" in text:
            text = text.replace("+", "")
        words = text.split()
        for word in words:
            if word.isdigit():
                self.founded_count = int(word)
                print("Metindeki sayı:", self.founded_count)

    def calculate_page_count(self):
        page_count = self.count // self.product_per_page
        if self.count <= self.founded_count:
            if self.count % self.product_per_page != 0:
                page_count += 1
        else:
            page_count = 1

        print("Sayfa sayısı:", page_count)
        return page_count

    def load_more_product(self):
        page_count = self.calculate_page_count()
        for i in range(1, page_count + 1):
            self.scrape_data(i)

    def scrape_data(self, page_count):
        try:
            self.browser.get(f"{self.search_url}&page={page_count}")
            element_scroll = self.wait.until(EC.presence_of_element_located((By.XPATH, consts.XPATH_LOAD_MORE_PRODUCT)))
            scroll_to_element(self.browser, element_scroll)
            products = self.wait.until(EC.presence_of_all_elements_located((By.XPATH, consts.XPATH_PRODUCTS)))
            self.format_products_list(products)
        except Exception as E:
            print(E)
            # raise exceptions.NoSuchElementException("Ürünler çekilirken hata")

    def format_products_list(self, products_list):
        for product_element in products_list:
            self._product(product_element)

    def get_product_details(self):
        for url in self.products_list:
            self.browser.get(url)

    def _product(self, product_element):
        image_list = product_element.find_elements(By.XPATH, './/img')
        images_list = []

        for image in image_list:
            images_list.append(image.get_attribute("src"))

        a_tag = product_element.find_element(By.XPATH, './/a')
        price = product_element.find_element(By.XPATH, ".//div[@data-test-id='price-current-price']").text

        product = {"name": a_tag.get_attribute("title"), "url": a_tag.get_attribute("href"),
                   "images": images_list,
                   "price": price}

        self.products_list.append(product)

    def write_to_excel(self):
        workbook = openpyxl.Workbook()

        # "Products" adında bir çalışma sayfası oluşturun
        worksheet = workbook.create_sheet('Products')

        header = ["Name", "URL", "Images", "Price"]

        # Başlıkları yazın
        for i, col_name in enumerate(header):
            worksheet.cell(row=1, column=i + 1, value=col_name)

        worksheet = workbook['Products']

        for i, p in enumerate(self.products_list):
            row = i + 2  # Başlık satırından sonra veri satırları başlar
            worksheet.cell(row=row, column=1, value=p['name'])
            worksheet.cell(row=row, column=2, value=p['url'])
            worksheet.cell(row=row, column=3, value=", ".join(p['images']))
            worksheet.cell(row=row, column=4, value=p['price'])

        # Excel dosyasını kaydedin
        workbook.save(f'{self.keyword.replace(" ", "_")}.xlsx')
